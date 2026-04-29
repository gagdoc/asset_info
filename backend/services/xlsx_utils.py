"""
xlsx_utils.py
=============
공용 Excel(.xlsx) 빌더 유틸리티.

대시보드 내보내기와 동일한 스타일(인디고 헤더, 얇은 테두리, 열 너비 자동)을
모든 내보내기 기능에 일관되게 적용합니다.
"""

from __future__ import annotations
import io
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 공용 스타일 상수 ───────────────────────────────────────────────────────────
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL   = PatternFill(fill_type="solid", fgColor="4F46E5")   # 인디고
_HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGN    = Alignment(vertical="center")
_THIN_SIDE     = Side(style="thin", color="D1D5DB")
_THIN_BORDER   = Border(
    left=_THIN_SIDE, right=_THIN_SIDE,
    top=_THIN_SIDE,  bottom=_THIN_SIDE,
)

# 짝수 행 배경 (연한 회색) — 가독성 향상
_ALT_FILL = PatternFill(fill_type="solid", fgColor="F9FAFB")

MAX_COL_WIDTH = 45   # 최대 열 너비 (글자 수 기준)
MIN_COL_WIDTH = 8    # 최소 열 너비


def _write_sheet(
    ws,
    columns: list[dict],   # [{"key": "item_name", "label": "품목명"}, ...]
    rows: list[dict],      # [{"item_name": "...", ...}, ...]
) -> None:
    """
    ws에 헤더 + 데이터 행을 기록하고 스타일을 적용합니다.

    columns:
        key   → rows 딕셔너리에서 값을 꺼내는 키 이름
        label → 헤더 셀에 표시되는 한국어 레이블
    """
    keys   = [c["key"]   for c in columns]
    labels = [c["label"] for c in columns]

    # 1. 헤더 행
    ws.append(labels)
    ws.row_dimensions[1].height = 24
    for col_idx, _ in enumerate(labels, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border    = _THIN_BORDER

    # 2. 데이터 행
    for row_num, row_data in enumerate(rows, start=2):
        values = [_fmt(row_data.get(k)) for k in keys]
        ws.append(values)
        use_alt = (row_num % 2 == 0)
        for col_idx, _ in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.alignment = _DATA_ALIGN
            cell.border    = _THIN_BORDER
            if use_alt:
                cell.fill = _ALT_FILL

    # 3. 열 너비 자동 조정
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col_cells),
            default=MIN_COL_WIDTH,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_len + 3, MIN_COL_WIDTH), MAX_COL_WIDTH
        )

    # 4. 행 높이 (데이터)
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 18

    # 5. 틀 고정 (헤더 행 아래)
    ws.freeze_panes = "A2"


def _fmt(value: Any) -> str | int | float:
    """셀 값 포맷팅 — None은 빈 문자열, 숫자는 숫자 타입 유지."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Y" if value else "N"
    if isinstance(value, (int, float)):
        return value
    return str(value)


def build_xlsx(
    sheets: list[dict],
    title: str = "",
) -> io.BytesIO:
    """
    여러 시트를 가진 xlsx 파일을 BytesIO로 반환합니다.

    sheets: [
        {
            "title":   "시트 탭 이름",
            "columns": [{"key": "...", "label": "..."}, ...],
            "rows":    [{...}, ...],
        },
        ...
    ]
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # 기본 빈 시트 제거

    if title:
        wb.properties.title = title

    for sheet_def in sheets:
        tab_title = sheet_def.get("title", "Sheet")[:31]   # Excel 탭명 31자 제한
        columns   = sheet_def.get("columns", [])
        rows      = sheet_def.get("rows", [])

        ws = wb.create_sheet(title=tab_title)

        if not columns or not rows:
            ws.append(["데이터 없음"])
            continue

        _write_sheet(ws, columns, rows)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
