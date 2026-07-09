import pandas as pd
from datetime import datetime
import io
import re
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from backend.services.database import (
    load_from_db, update_db, normalize_email, get_connection, ASSET_DB_FILE
)

def get_dashboard_integrated_data() -> List[Dict[str, Any]]:
    dfs = load_from_db()
    if "All_User" not in dfs or dfs["All_User"].empty:
        return []

    view_df = dfs["All_User"].copy()
    
    # Drop known asset columns from All_User to avoid merge conflicts (_x/_y)
    asset_cols_to_drop = ["Lease_List", "Ipad_List", "TeamsNum", "Printer", "Monitor", "모니터"]
    cols_present = [c for c in asset_cols_to_drop if c in view_df.columns]
    if cols_present:
        view_df.drop(columns=cols_present, inplace=True)

    # Filter valid emails
    if "email" in view_df.columns:
        view_df["email"] = view_df["email"].astype(str).str.strip()
        view_df = view_df[~view_df["email"].str.lower().isin(["nan", "", "none", "null"])]
        view_df["_email_key"] = view_df["email"].str.lower()
    else:
        return []

    # Helper: Group by email and join values with duplicate marker
    def _group_asset(df, email_col, val_col, target_key):
        if df.empty or email_col not in df.columns or val_col not in df.columns:
            return pd.DataFrame(columns=["_email_key", target_key])
        
        subset = df[[email_col, val_col]].dropna().copy()
        subset["_email_key"] = subset[email_col].astype(str).str.strip().str.lower()
        subset[val_col] = subset[val_col].astype(str).str.strip()
        
        # Filter out invalid values
        subset = subset[~subset[val_col].isin(["", "-", "nan", "None", "null"])]
        if subset.empty:
            return pd.DataFrame(columns=["_email_key", target_key])

        # Group and Join
        def _join_logic(x):
            unique_vals = sorted(list(set(x)))
            if not unique_vals: return "-"
            prefix = "[중복!] " if len(unique_vals) > 1 else ""
            return prefix + ", ".join(unique_vals)

        grouped = subset.groupby("_email_key")[val_col].apply(_join_logic).reset_index()
        return grouped.rename(columns={val_col: target_key})

    # 1. Lease
    if "Lease" in dfs and not dfs["Lease"].empty:
        cols = dfs["Lease"].columns
        target_col = "S/N" if "S/N" in cols else (cols[3] if len(cols) > 3 else cols[0])
        l_sub = _group_asset(dfs["Lease"], "email", target_col, "Lease_List")
        view_df = pd.merge(view_df, l_sub, on="_email_key", how="left")

    # 2. iPad
    if "iPad" in dfs and not dfs["iPad"].empty:
        cols = dfs["iPad"].columns
        target_col = "S/N" if "S/N" in cols else ("Model" if "Model" in cols else cols[0])
        i_sub = _group_asset(dfs["iPad"], "email", target_col, "Ipad_List")
        view_df = pd.merge(view_df, i_sub, on="_email_key", how="left")

    # 3. Monitor
    if "Monitor" in dfs and not dfs["Monitor"].empty:
        m_sub = _group_asset(dfs["Monitor"], "email", "Model", "Monitor")
        view_df = pd.merge(view_df, m_sub, on="_email_key", how="left")

    # 4. Teams
    if "Teams" in dfs and not dfs["Teams"].empty:
        cols = dfs["Teams"].columns
        target_col = next((c for c in ["TeamsNumber", "Number", "전화번호"] if c in cols), None)
        if target_col:
            t_sub = _group_asset(dfs["Teams"], "email", target_col, "TeamsNum")
            view_df = pd.merge(view_df, t_sub, on="_email_key", how="left")

    # 5. Printer
    if "Printer" in dfs and not dfs["Printer"].empty:
        cols = dfs["Printer"].columns
        target_col = next((c for c in ["Additional Information 2", "프린터정보", "Model"] if c in cols), None)
        if target_col:
            p_sub = _group_asset(dfs["Printer"], "email", target_col, "Printer")
            view_df = pd.merge(view_df, p_sub, on="_email_key", how="left")

    # 6. Resign Info
    if "Resign" in dfs and not dfs["Resign"].empty and "email" in dfs["Resign"].columns:
        try:
            r_sub = dfs["Resign"].copy()
            r_sub["_email_key"] = r_sub["email"].astype(str).str.strip().str.lower()
            
            if all(c in r_sub.columns for c in ["년도", "월", "날짜"]):
                r_sub["퇴사정보"] = r_sub.apply(lambda x: f"{int(x['년도'])}년 {int(x['월'])}월 {int(x['날짜'])}일 퇴사" if pd.notnull(x["년도"]) else "퇴사자", axis=1)
            elif "월" in r_sub.columns and "날짜" in r_sub.columns:
                r_sub["퇴사정보"] = r_sub.apply(lambda x: f"{int(x['월'])}월 {int(x['날짜'])}일 퇴사" if pd.notnull(x["월"]) else "퇴사자", axis=1)
            else:
                r_sub["퇴사정보"] = "퇴사자 목록 포함"
            
            r_sub = r_sub[["_email_key", "퇴사정보"]].drop_duplicates("_email_key")
            view_df = pd.merge(view_df, r_sub, on="_email_key", how="left")
        except: pass

    # 임시 병합 키 제거
    if "_email_key" in view_df.columns:
        view_df.drop(columns=["_email_key"], inplace=True)

    # Fill missing cols and handle NaNs
    view_df["퇴사정보"] = view_df.get("퇴사정보", pd.Series()).fillna("-")

    if "모니터" in view_df.columns and "Monitor" not in view_df.columns:
        view_df.rename(columns={"모니터": "Monitor"}, inplace=True)

    desired_cols = ["NAME", "이름", "email", "BU", "ROLE", "Lease_List", "Ipad_List", "TeamsNum", "Printer", "Monitor", "퇴사정보"]
    
    # Fill missing cols
    for col in desired_cols:
        if col not in view_df.columns:
            view_df[col] = "-"
        else:
            view_df[col] = view_df[col].fillna("-")
            
    # Final sort/filter
    result_df = view_df[desired_cols].where(pd.notnull(view_df[desired_cols]), "-")
    return result_df.to_dict(orient="records")

def perform_bulk_search(search_input: str, search_type: str, search_target: str) -> Dict[str, Any]:
    terms = [t.strip() for t in re.split(r'[\n,;]+', search_input) if t.strip()]
    if not terms:
        return {"found": [], "notFound": []}

    found_results = []
    found_terms = set()

    def filter_columns_by_type(all_cols, table_key):
        if search_type == "email":
            return [c for c in all_cols if c.lower() in ["email"]]
        elif search_type == "serial_laptop":
            if table_key == "Lease":
                return [c for c in all_cols if c in ["S/N", "SNOW Tag"]]
            elif table_key == "dashboard":
                return [c for c in all_cols if c == "Lease_List"]
            return []
        elif search_type == "serial_ipad":
            if table_key == "iPad":
                return [c for c in all_cols if c in ["S/N", "전화 번호", "Model"]]
            elif table_key == "dashboard":
                return [c for c in all_cols if c == "Ipad_List"]
            return []
        return all_cols

    if search_target == "dashboard":
        dashboard_data = get_dashboard_integrated_data()
        if dashboard_data:
            df = pd.DataFrame(dashboard_data)
            base_cols = df.columns.tolist()
            search_cols = filter_columns_by_type(base_cols, "dashboard")
            
            for term in terms:
                term_lower = term.lower()
                matched_for_term = False
                if search_cols:
                    for col in search_cols:
                        mask = df[col].astype(str).str.lower().str.contains(term_lower, na=False, regex=False)
                        if mask.any():
                            matched_rows = df[mask].copy()
                            for _, row in matched_rows.iterrows():
                                row_dict = row.where(pd.notnull(row), None).to_dict()
                                res_entry = {
                                    "type": "대시보드",
                                    "table_key": "dashboard",
                                    "match_col": col,
                                    "match_term": term,
                                    "data": row_dict
                                }
                                if not any(r["data"] == row_dict and r["match_term"] == term for r in found_results):
                                    found_results.append(res_entry)
                                matched_for_term = True
                if matched_for_term:
                    found_terms.add(term)
    else:
        dfs = load_from_db()
        search_targets = {
            "Lease": ["S/N", "SNOW Tag", "email", "User", "Model"],
            "iPad": ["S/N", "전화 번호", "email", "User", "Model"],
            "Teams": ["TeamsNumber", "Number", "email", "History"],
            "Monitor": ["Model", "email", "User"],
            "Printer": ["Model", "email", "프린터정보"],
            "All_User": ["email", "NAME", "이름"]
        }
        type_labels = {
            "Lease": "노트북", "iPad": "아이패드", "Teams": "Teams", "Monitor": "모니터", "Printer": "복합기", "All_User": "대시보드"
        }
        
        target_keys = list(search_targets.keys())
        if search_target == "laptop":
            target_keys = ["Lease"]
        elif search_target == "ipad":
            target_keys = ["iPad"]

        for term in terms:
            term_lower = term.lower()
            matched_for_term = False
            for table_key in target_keys:
                if table_key not in dfs or dfs[table_key].empty:
                    continue
                df = dfs[table_key]
                available_cols = [c for c in search_targets[table_key] if c in df.columns]
                search_cols = filter_columns_by_type(available_cols, table_key)
                if not search_cols: continue
                for col in search_cols:
                    mask = df[col].astype(str).str.lower().str.contains(term_lower, na=False, regex=False)
                    if mask.any():
                        matched_rows = df[mask].copy()
                        for _, row in matched_rows.iterrows():
                            row_dict = row.where(pd.notnull(row), None).to_dict()
                            res_entry = {
                                "type": type_labels.get(table_key, table_key),
                                "table_key": table_key,
                                "match_col": col,
                                "match_term": term,
                                "data": row_dict
                            }
                            found_results.append(res_entry)
                            matched_for_term = True
            if matched_for_term:
                found_terms.add(term)

    return {"found": found_results, "notFound": list(terms[i] for i in range(len(terms)) if terms[i] not in found_terms)}

def enrich_data_with_assets(target_df: pd.DataFrame, dfs: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    target_df = normalize_email(target_df)
    asset_map = {"NAME": {}, "BU": {}, "노트북": {}, "아이패드": {}, "모니터": {}, "Teams": {}, "복합기": {}}
    
    def robust_to_dict(df_src, email_col, val_col):
        if df_src.empty or email_col not in df_src.columns or val_col not in df_src.columns:
            return {}
        df_clean = df_src[[email_col, val_col]].dropna().copy()
        df_clean[email_col] = df_clean[email_col].astype(str).str.strip().str.lower()
        df_clean[val_col] = df_clean[val_col].astype(str).str.strip().str.lstrip("=")
        placeholders = ["", "-", "nan", "None", ".", "0", "없음"]
        df_clean = df_clean[~df_clean[val_col].isin(placeholders)]
        if df_clean.empty: return {}
        return df_clean.groupby(email_col)[val_col].apply(lambda x: ", ".join(sorted(set(x)))).to_dict()
    
    if "All_User" in dfs and not dfs["All_User"].empty:
        df_users = dfs["All_User"].dropna(subset=["email"])
        if "NAME" in df_users.columns:
            asset_map["NAME"] = {str(k).strip().lower(): v for k, v in df_users.set_index("email")["NAME"].to_dict().items()}
        if "BU" in df_users.columns:
            asset_map["BU"] = {str(k).strip().lower(): v for k, v in df_users.set_index("email")["BU"].to_dict().items()}
    
    # ... rest of enrich logic ...
    for key, label, search_cols in [
        ("Lease", "노트북", ["S/N"]), ("iPad", "아이패드", ["S/N", "Model"]),
        ("Monitor", "모니터", ["Model"]), ("Teams", "Teams", ["TeamsNumber", "Number", "전화번호"]),
        ("Printer", "복합기", ["Model", "프린터정보"])
    ]:
        if key in dfs and not dfs[key].empty:
            cols = dfs[key].columns
            target_col = next((c for c in search_cols if c in cols), (cols[3] if len(cols) > 3 else cols[0]))
            asset_map[label].update(robust_to_dict(dfs[key], "email", target_col))

    enrich_cols = {"NAME": "NAME", "BU": "BU", "노트북": "노트북", "아이패드": "아이패드", "모니터": "모니터", "Teams": "Teams", "복합기": "복합기"}
    for col in enrich_cols.keys():
        if col not in target_df.columns: target_df[col] = None
    
    for idx, row in target_df.iterrows():
        email = str(row.get("email", "")).strip().lower()
        if not email: continue
        for df_col, map_key in enrich_cols.items():
            val = target_df.at[idx, df_col]
            new_val = asset_map[map_key].get(email)
            if map_key in ["노트북", "아이패드", "모니터", "Teams", "복합기"]:
                target_df.at[idx, df_col] = str(new_val) if new_val else "-"
            else:
                if pd.isna(val) or str(val).strip() in ["", "-", "nan", "None", ".", "0"]:
                    target_df.at[idx, df_col] = str(new_val) if new_val else "-"
    
    target_df["_is_deleted"] = True
    if "All_User" in dfs and not dfs["All_User"].empty and "email" in dfs["All_User"].columns:
        master_emails = set(dfs["All_User"]["email"].dropna().astype(str).str.strip().str.lower())
        if "email" in target_df.columns:
            target_df["_is_deleted"] = ~target_df["email"].astype(str).str.strip().str.lower().isin(master_emails)
    
    target_df["is_resigned"] = False
    if "Resign" in dfs and not dfs["Resign"].empty and "email" in dfs["Resign"].columns:
        resigned_emails = set(dfs["Resign"]["email"].dropna().astype(str).str.strip().str.lower())
        if "email" in target_df.columns:
            target_df["is_resigned"] = target_df["email"].astype(str).str.strip().str.lower().isin(resigned_emails)
            
    return target_df

def return_asset(email: str, asset_type: Optional[str], name: Optional[str], bu: Optional[str]) -> Dict[str, Any]:
    today_str = datetime.now().strftime("%Y%m%d")
    dfs = load_from_db()
    updated_tables = []
    
    eng_name = name or "Unknown"
    user_bu = bu or "Unknown"
    
    if (eng_name == "Unknown" or user_bu == "Unknown") and "All_User" in dfs and not dfs["All_User"].empty:
        clean_email = email.split('@')[0] if '@' in email else email
        mask = dfs["All_User"]["email"].str.strip().str.lower().str.contains(clean_email, na=False, regex=False)
        user_row = dfs["All_User"][mask]
        if not user_row.empty:
            if eng_name == "Unknown": eng_name = str(user_row.iloc[0].get("NAME", "Unknown")).strip()
            if user_bu == "Unknown": user_bu = str(user_row.iloc[0].get("BU", "Unknown")).strip()
    
    if eng_name == "Unknown": eng_name = email.split('@')[0]
    teams_label = f"{user_bu}/{eng_name}" if user_bu != "Unknown" else eng_name
    
    def process_table(table_key, label, extra_updates=None):
        if table_key in dfs and not dfs[table_key].empty and "email" in dfs[table_key].columns:
            mask = dfs[table_key]["email"].str.strip().str.lower() == email
            if mask.any():
                dfs[table_key].loc[mask, "email"] = ""
                if extra_updates:
                    for col, val in extra_updates.items():
                        if col in dfs[table_key].columns: dfs[table_key].loc[mask, col] = val
                update_db(table_key, dfs[table_key])
                return label
        return None

    mapping = {
        "Lease": ("Lease", "PC/노트북", {"User": "STOCK", "BU": "IT", "Additional Information": f"{today_str}/{email}/반납"}),
        "iPad": ("iPad", "아이패드", {"User": "STOCK", "BU": "IT", "Role": "IT", "Additional Information": f"{today_str}/{email}/반납"}),
        "Teams": ("Teams", "팀즈 번호", {"History": f"{today_str}/{email}/반납 {teams_label}"}),
        "Monitor": ("Monitor", "모니터", {"Additional Information": f"{today_str}/{email}/반납"}),
        "Printer": ("Printer", "복합기", {"Additional Information": f"{today_str}/{email}/반납"})
    }

    if asset_type:
        if asset_type in mapping:
            res = process_table(*mapping[asset_type])
            if res: updated_tables.append(res)
    else:
        for k in mapping:
            res = process_table(*mapping[k])
            if res: updated_tables.append(res)

    # Clear Resign table
    if "Resign" in dfs and not dfs["Resign"].empty and "email" in dfs["Resign"].columns:
        mask = dfs["Resign"]["email"].str.strip().str.lower() == email
        if mask.any():
            cols = ["노트북", "아이패드", "모니터", "복합기", "Teams"] if not asset_type else [asset_type]
            for col in cols:
                if col in dfs["Resign"].columns: dfs["Resign"].loc[mask, col] = "-"
            update_db("Resign", dfs["Resign"])

    if updated_tables:
        return {"success": True, "message": f"반납 처리 완료: {', '.join(updated_tables)}"}
    return {"success": False, "message": "할당된 자산이 없거나 이미 반납되었습니다."}


# ─── 전체 데이터 Excel 내보내기 ─────────────────────────────────

# 선택 가능한 시트 정의
EXPORT_SHEET_MAP = {
    "Lease":     "노트북",
    "iPad":      "아이패드",
    "Monitor":   "모니터",
    "Printer":   "프린터",
    "Teams":     "Teams번호",
    "NewHire":   "신규입사자",
    "Resign":    "퇴사자",
    "Dashboard": "자산통합현황",
}

def export_sheets_to_excel(sheet_keys: List[str]) -> io.BytesIO:
    """
    요청된 sheet_keys 에 해당하는 데이터를 Google Sheets에서 읽어
    각각 별도 탭으로 구성된 단일 .xlsx 파일을 BytesIO로 반환한다.
    """
    dfs = load_from_db()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(fill_type="solid", fgColor="4F46E5")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side    = Side(style="thin", color="CCCCCC")
    thin_border  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    valid_keys = [k for k in sheet_keys if k in EXPORT_SHEET_MAP]
    if not valid_keys:
        valid_keys = list(EXPORT_SHEET_MAP.keys())

    for key in valid_keys:
        sheet_title = EXPORT_SHEET_MAP[key]

        # ── Dashboard(자산 통합 상세 조회) 특별 처리 ──────────
        if key == "Dashboard":
            raw = get_dashboard_integrated_data()
            df = pd.DataFrame(raw) if raw else pd.DataFrame()
            # 컬럼명을 한국어 레이블로 변환
            col_rename = {
                "NAME":       "이름(영문)",
                "이름":        "이름(한글)",
                "email":      "이메일",
                "BU":         "부서",
                "ROLE":       "직책",
                "Lease_List": "노트북 S/N",
                "Ipad_List":  "아이패드 S/N",
                "TeamsNum":   "Teams 번호",
                "Printer":    "복합기",
                "Monitor":    "모니터",
                "퇴사정보":     "비고",
            }
            df = df.rename(columns={k: v for k, v in col_rename.items() if k in df.columns})
        else:
            df = dfs.get(key, pd.DataFrame())

        # 불필요한 내부 컬럼 제거
        df = df.copy()
        drop_cols = [c for c in df.columns if c.startswith("Unnamed")]
        df.drop(columns=drop_cols, inplace=True, errors="ignore")

        ws = wb.create_sheet(title=sheet_title)

        if df.empty:
            ws.append(["데이터 없음"])
            continue

        # 헤더 행
        headers = df.columns.tolist()
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 데이터 행
        for row_data in df.fillna("").values.tolist():
            ws.append([str(v) if v is not None else "" for v in row_data])

        # 테두리 + 셀 정렬 (데이터 행)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border    = thin_border
                cell.alignment = Alignment(vertical="center")

        # 컬럼 너비 자동 조정
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max(
                (len(str(cell.value)) if cell.value else 0 for cell in col_cells),
                default=8
            )
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = min(max_len + 4, 40)

        # 헤더 행 높이
        ws.row_dimensions[1].height = 22

        # 1행(헤더) 고정 및 필터 적용
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
