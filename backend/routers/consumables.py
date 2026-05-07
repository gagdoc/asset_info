from fastapi import APIRouter, HTTPException, Query, Body, Path
from fastapi.responses import StreamingResponse
from backend.services.consumables_service import (
    get_available_months, get_items_list, get_outbound_history,
    get_estimate, add_outbound, save_item, update_outbound_history,
    delete_outbound_history, get_item_outbound_history,
    create_month_sheet, invalidate_cache, get_tonner_consignment_history,
    get_toner_inventory, update_toner_item, delete_item,
    get_inbound_history, add_inbound, delete_inbound, update_inbound,
    get_inventory_report, sync_toner_to_items_list,
    get_month_close_status, confirm_month_snapshot, close_month,
    reopen_month, get_monthly_toner_report, reset_month_snapshot,
    get_purchase_history, add_purchase_record, delete_purchase_record,
    set_toner_stock_direct,
    get_individual_inbound_history, add_individual_inbound, delete_individual_inbound,
    IS_PRODUCTION,
)
from typing import List, Dict, Any
import io
import os
from datetime import datetime

router = APIRouter(
    prefix="/api/consumables",
    tags=["consumables"],
    responses={404: {"description": "Not found"}},
)

@router.get("/clear-cache")
def clear_consumables_cache():
    """서버 캐시를 강제로 비워 구글 시트 최신 데이터를 불러올 수 있게 함"""
    invalidate_cache()
    return {"status": "success", "message": "Cache cleared."}

@router.get("/months")
def list_months():
    """사용 가능한 월(시트) 목록 반환"""
    months = get_available_months()
    return {"months": months}

@router.post("/months")
def add_new_month(data: Dict[str, str] = Body(...)):
    """새로운 월의 출고 내역 시트 생성"""
    month_name = data.get("month")
    start_date = data.get("start_date", "")
    if not month_name:
        raise HTTPException(status_code=400, detail="새로운 월 이름이 필요합니다.")
        
    success = create_month_sheet(month_name, start_date)
    if not success:
        raise HTTPException(status_code=500, detail="이미 존재하는 월이거나 구글 시트 생성에 실패했습니다.")
    return {"status": "success"}

@router.get("/items")
def list_items(
    month: str = Query(None, description="월 필터 (예: '2026년 4월')"),
    dispatch_mode: str = Query("cumulative", description="출고 집계 모드: 'cumulative'(누적) | 'monthly'(월별)")
):
    """품목 리스트 반환. dispatch_mode=monthly 시 지정 month의 출고만 집계."""
    items = get_items_list(month=month, dispatch_mode=dispatch_mode)
    return items

@router.get("/items/{item_name}/outbound")
def list_item_outbounds(item_name: str = Path(..., description="조회할 품목 이름")):
    """선택한 단일 품목의 전체 시트(월) 기준 과거 출고 이력 조회"""
    history = get_item_outbound_history(item_name)
    return history

@router.get("/outbound")
def list_outbound(month: str = Query(..., description="조회할 월 (예: '3월')")):
    """선택한 월의 출고 이력 반환"""
    history = get_outbound_history(month)
    return history

@router.get("/estimate")
def get_month_estimate(month: str = Query(..., description="조회할 월 (예: '3월')")):
    """선택한 월의 견적서 데이터 반환"""
    estimate = get_estimate(month)
    return estimate

@router.get("/estimate/download")
def download_estimate_excel(month: str = Query(..., description="다운로드할 월 (예: '3월')")):
    """선택한 월의 견적서를 Excel 파일로 다운로드 (template 없이 직접 생성)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 패키지가 필요합니다.")

    try:
        estimate = get_estimate(month)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"견적 데이터 조회 오류: {str(e)}")

    if not estimate:
        raise HTTPException(status_code=404, detail=f"{month} 견적 데이터가 없습니다.")

    today_str = datetime.now().strftime("%Y%m%d")
    today_fmt = datetime.now().strftime("%Y-%m-%d")

    # ── 스타일 헬퍼 ────────────────────────────────────────────────
    def thin_border(top=False, bottom=False, left=False, right=False):
        s = Side(border_style="thin", color="000000")
        return Border(
            top=s if top else Side(),
            bottom=s if bottom else Side(),
            left=s if left else Side(),
            right=s if right else Side(),
        )

    def set_cell(ws, coord, value, bold=False, size=10, align="left",
                 valign="center", num_fmt=None, fill=None, border=None):
        c = ws[coord]
        c.value = value
        c.font = Font(bold=bold, size=size, name="맑은 고딕")
        c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=True)
        if num_fmt:
            c.number_format = num_fmt
        if fill:
            c.fill = fill
        if border:
            c.border = border

    # ── 워크북 생성 ────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"견적서_{month}"

    # 열 너비 설정 (A:숨김, B:NO, C:ITEM, D:품목명, E:수량, F:단가, G:금액)
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16

    # ── 헤더 영역 ──────────────────────────────────────────────────
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_font = Font(bold=True, size=16, color="FFFFFF", name="맑은 고딕")

    # 제목
    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = "소 모 품 견 적 서"
    c.font = title_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = header_fill
    ws.row_dimensions[2].height = 36

    # 날짜 / 견적번호
    ws.row_dimensions[3].height = 18
    set_cell(ws, "B3", "발행일자", bold=True, size=9, align="center")
    ws.merge_cells("C3:D3")
    set_cell(ws, "C3", today_fmt, size=9)
    set_cell(ws, "F3", "No.", bold=True, size=9, align="center")
    month_num = ''.join(filter(str.isdigit, month))
    set_cell(ws, "G3", f"{today_str}-{month_num}", size=9, align="center")

    # 월 정보
    ws.row_dimensions[4].height = 18
    ws.merge_cells("B4:G4")
    set_cell(ws, "B4", f"조회 기준: {month}", size=9, align="right")

    ws.row_dimensions[5].height = 6  # 여백

    # ── 테이블 헤더 ────────────────────────────────────────────────
    col_header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    col_header_font = Font(bold=True, size=10, color="FFFFFF", name="맑은 고딕")

    HEADER_ROW = 6
    ws.row_dimensions[HEADER_ROW].height = 22
    headers = [("B", "No.", "center"), ("C", "ITEM (분류)", "center"),
               ("D", "품목명 (Model Name)", "center"), ("E", "수량", "center"),
               ("F", "단가 (₩)", "center"), ("G", "견적금액 (₩)", "center")]
    for col, label, align in headers:
        c = ws[f"{col}{HEADER_ROW}"]
        c.value = label
        c.font = col_header_font
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.fill = col_header_fill
        c.border = thin_border(top=True, bottom=True,
                               left=(col == "B"), right=(col == "G"))

    # ── 데이터 행 ──────────────────────────────────────────────────
    DATA_START = HEADER_ROW + 1
    n = len(estimate)
    row_fills = [
        PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid"),
        PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    ]

    def parse_int(val):
        """문자열 숫자("1,500" 또는 "3" 등)를 정수로 안전하게 변환"""
        try:
            return int(str(val).replace(',', '').strip() or 0)
        except (ValueError, TypeError):
            return 0

    for idx, row in enumerate(estimate):
        r = DATA_START + idx
        ws.row_dimensions[r].height = 18
        fill = row_fills[idx % 2]
        bdr_top = (idx == 0)
        bdr_bot = (idx == n - 1)

        # get_estimate()는 수치를 문자열로 반환하므로 정수로 변환
        qty_val   = parse_int(row.get("total_qty", 0))
        unit_val  = parse_int(row.get("unit_price", 0))
        total_p   = parse_int(row.get("total_price", 0))

        data_cols = [
            ("B", row.get("no", idx+1), False, "center", None),
            ("C", row.get("category", ""), False, "center", None),
            ("D", row.get("item_name", ""), True, "left", None),
            ("E", qty_val,  True,  "center", None),
            ("F", unit_val, False, "right",  '#,##0'),
            ("G", total_p,  True,  "right",  '#,##0'),
        ]
        for col, val, bold, align, nfmt in data_cols:
            c = ws[f"{col}{r}"]
            c.value = val
            c.font = Font(bold=bold, size=10, name="맑은 고딕")
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.fill = fill
            c.border = thin_border(top=bdr_top, bottom=bdr_bot,
                                   left=(col == "B"), right=(col == "G"))
            if nfmt:
                c.number_format = nfmt

    # ── TOTAL 행 ───────────────────────────────────────────────────
    TOTAL_ROW = DATA_START + n
    ws.row_dimensions[TOTAL_ROW].height = 22
    total_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    ws.merge_cells(f"B{TOTAL_ROW}:F{TOTAL_ROW}")
    c = ws[f"B{TOTAL_ROW}"]
    c.value = "TOTAL AMOUNT (VAT 별도)"
    c.font = Font(bold=True, size=10, color="FFFFFF", name="맑은 고딕")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = total_fill
    c.border = thin_border(top=True, bottom=True, left=True)

    # total_price도 문자열이므로 정수 변환 후 합산
    total_val = sum(parse_int(r.get("total_price", 0)) for r in estimate)
    c = ws[f"G{TOTAL_ROW}"]
    c.value = total_val
    c.font = Font(bold=True, size=11, color="FFFFFF", name="맑은 고딕")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.fill = total_fill
    c.border = thin_border(top=True, bottom=True, right=True)
    c.number_format = '#,##0'

    # ── VAT 포함 합계 안내 ─────────────────────────────────────────
    NOTE_ROW = TOTAL_ROW + 1
    ws.row_dimensions[NOTE_ROW].height = 16
    ws.merge_cells(f"B{NOTE_ROW}:G{NOTE_ROW}")
    set_cell(ws, f"B{NOTE_ROW}",
             f"  ※ 상기 금액은 부가세(VAT 10%) 별도 금액입니다.  |  VAT 포함: {int(total_val * 1.1):,} ₩",
             size=9, align="right")

    # ── 메모리 스트림으로 반환 ─────────────────────────────────────
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"견적서_{month}_{today_str}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

@router.post("/outbound")
def create_outbound(data: Dict[str, Any] = Body(...)):
    """월별 출고 데이터 추가 (마감된 월은 차단)"""
    month = data.get("month")
    if not month:
        raise HTTPException(status_code=400, detail="Month is required")

    # 마감된 월에는 신규 출고 불가
    status_info = get_month_close_status(month)
    if status_info["status"] == "closed":
        raise HTTPException(status_code=403, detail=f"'{month}'은(는) 마감된 월입니다. 출고 내역을 추가할 수 없습니다.")

    success = add_outbound(month, data)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to add outbound record")
    return {"status": "success"}


# ── 월별 재고 마감 엔드포인트 ───────────────────────────────────

@router.get("/month-status")
def get_month_status(month: str = Query(..., description="조회할 월 (예: '2026년4월')")):
    """월의 마감 상태 조회: open / confirmed / closed"""
    return get_month_close_status(month)


@router.post("/month-confirm")
def confirm_snapshot(data: Dict[str, Any] = Body(...)):
    """이달 재고 확정: 현재 토너 재고를 해당 월 시작 재고로 스냅샷 저장"""
    month = data.get("month")
    if not month:
        raise HTTPException(status_code=400, detail="month는 필수입니다.")
    result = confirm_month_snapshot(month)
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "스냅샷 저장 실패"))
    return result


@router.post("/month-close")
def close_month_endpoint(data: Dict[str, Any] = Body(...)):
    """월 마감: 신규 출고 추가 차단 (수정은 허용)"""
    month = data.get("month")
    if not month:
        raise HTTPException(status_code=400, detail="month는 필수입니다.")
    result = close_month(month)
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "마감 실패"))
    return result


@router.post("/month-reopen")
def reopen_month_endpoint(data: Dict[str, Any] = Body(...)):
    """마감 해제: closed → confirmed (수정을 위한 임시 해제)"""
    month = data.get("month")
    if not month:
        raise HTTPException(status_code=400, detail="month는 필수입니다.")
    result = reopen_month(month)
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "마감 해제 실패"))
    return result


@router.get("/monthly-report")
def get_monthly_report(month: str = Query(..., description="보고서 조회 월 (예: '2026년4월')")):
    """월별 토너 재고 보고서: 시작재고 / 출고량 / 잔여재고"""
    return get_monthly_toner_report(month)


@router.get("/app-env")
def get_app_env():
    """현재 실행 환경 반환 (프론트엔드 개발 모드 배너용)"""
    return {"is_production": IS_PRODUCTION, "app_env": "production" if IS_PRODUCTION else "development"}


@router.post("/month-reset")
def reset_month(data: Dict[str, Any] = Body(...)):
    """[개발 전용] 월 스냅샷/마감 상태 초기화 → open으로 리셋. 프로덕션에서는 차단."""
    month = data.get("month")
    if not month:
        raise HTTPException(status_code=400, detail="month는 필수입니다.")
    result = reset_month_snapshot(month)
    if not result.get("success"):
        raise HTTPException(status_code=403, detail=result.get("error", "초기화 실패"))
    return result

# ── 구매 입고 내역 ─────────────────────────────────────────────

@router.get("/purchase-history")
def list_purchase_history():
    """구매 입고 내역 전체 조회"""
    return get_purchase_history()

@router.post("/purchase-history")
def create_purchase_record(data: Dict[str, Any] = Body(...)):
    """구매 입고 내역 추가"""
    if not data.get("item_name"):
        raise HTTPException(status_code=400, detail="품목명은 필수입니다.")
    if not data.get("date"):
        raise HTTPException(status_code=400, detail="날짜는 필수입니다.")
    result = add_purchase_record(data)
    if not result.get("success"):
        raise HTTPException(status_code=500, detail=result.get("error", "등록 실패"))
    return {"status": "success"}

@router.delete("/purchase-history/{row_index}")
def remove_purchase_record(row_index: int = Path(..., description="삭제할 행 번호")):
    """구매 입고 내역 삭제"""
    result = delete_purchase_record(row_index)
    if not result.get("success"):
        raise HTTPException(status_code=500, detail=result.get("error", "삭제 실패"))
    return {"status": "success"}


@router.post("/items")
def create_or_update_item(data: Dict[str, Any] = Body(...)):
    """품목 마스터 리스트 추가/수정"""
    if not data.get("item_name"):
        raise HTTPException(status_code=400, detail="Item name is required")
        
    success = save_item(data)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to save item")
    return {"status": "success"}

@router.delete("/items")
def delete_item_endpoint(
    row_index: int = Query(..., description="삭제할 행 번호"),
    item_name: str = Query(..., description="삭제할 품목명 (이중 검증용)")
):
    """품목 마스터 리스트에서 특정 품목 삭제 (row_index + item_name 이중 검증)"""
    success = delete_item(row_index, item_name)
    if not success:
        raise HTTPException(status_code=500, detail="품목 삭제에 실패했습니다. 행 정보를 확인하세요.")
    return {"status": "success"}

@router.put("/outbound")
def update_outbound(data: Dict[str, Any] = Body(...)):
    """월별 출고 개별 데이터 수정"""
    month = data.get("month")
    row_index = data.get("row_index")
    if not month or not row_index:
        raise HTTPException(status_code=400, detail="Month and row_index are required")
        
    success = update_outbound_history(month, int(row_index), data)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to update outbound record")
    return {"status": "success"}

@router.delete("/outbound")
def delete_outbound(
    month: str = Query(...),
    row_index: int = Query(...),
    verify_date: str = Query("", description="삭제 전 날짜 검증값"),
    verify_item: str = Query("", description="삭제 전 품목명 검증값"),
    verify_user: str = Query("", description="삭제 전 사용자명 검증값 (중복 행 오탐 방지)"),
):
    """월별 출고 개별 데이터 삭제 (날짜+품목+사용자 3중 검증)"""
    success = delete_outbound_history(month, row_index, verify_date, verify_item, verify_user)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to delete outbound record")
    return {"status": "success"}

@router.get("/tonner-consignment")
def list_tonner_consignment(month: str = Query(None, description="조회할 월 (없으면 전체)")):
    """위탁 출고된 Tonner 내역 조회"""
    history = get_tonner_consignment_history(month=month)
    return history

@router.post("/sync-toner")
def sync_toner():
    """토너 재고 시트 → 품목리스트 단방향 동기화.
    토너 재고 시트에 있는 품목 중 품목리스트에 없는 항목을 자동으로 추가합니다."""
    result = sync_toner_to_items_list()
    return {
        "success": True,
        "added_count": len(result["added"]),
        "added": result["added"],
        "skipped_count": len(result["skipped"]),
    }

@router.get("/toner-inventory")
def list_toner_inventory():
    """토너 전용 재고 시트 전체 조회 (헤더 + 품목 목록)"""
    return get_toner_inventory()

@router.put("/toner-inventory")
def update_toner_inventory_item(data: Dict[str, Any] = Body(...)):
    """토너 재고 시트의 특정 행 수정"""
    row_index = data.get("row_index")
    if not row_index:
        raise HTTPException(status_code=400, detail="row_index is required")
    success = update_toner_item(int(row_index), data)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to update toner inventory item")
    return {"status": "success"}


# ─── 입고 이력 / 재고 입출고 현황 ──────────────────────────────

@router.get("/inbound")
def list_inbound(item_name: str = Query(None, description="특정 품목 필터 (없으면 전체)")):
    """입고 이력 조회 (전체 또는 품목별)"""
    return get_inbound_history(item_name=item_name)


@router.post("/inbound")
def create_inbound(data: Dict[str, Any] = Body(...)):
    """입고 기록 추가"""
    date = data.get("date", "")
    item_name = data.get("item_name", "")
    quantity = data.get("quantity", 0)
    memo = data.get("memo", "")
    if not date or not item_name:
        raise HTTPException(status_code=400, detail="date와 item_name은 필수입니다.")
    try:
        qty = int(str(quantity).replace(",", ""))
    except ValueError:
        raise HTTPException(status_code=400, detail="quantity는 숫자여야 합니다.")
    success = add_inbound(date, item_name, qty, memo)
    if not success:
        raise HTTPException(status_code=500, detail="입고 기록 추가에 실패했습니다.")
    return {"status": "success"}


@router.put("/inbound")
def modify_inbound(data: Dict[str, Any] = Body(...)):
    """입고 기록 수정"""
    row_index = data.get("row_index")
    date = data.get("date", "")
    item_name = data.get("item_name", "")
    quantity = data.get("quantity", 0)
    memo = data.get("memo", "")
    verify_item = data.get("verify_item", item_name)
    if not row_index or not date or not item_name:
        raise HTTPException(status_code=400, detail="row_index, date, item_name은 필수입니다.")
    try:
        qty = int(str(quantity).replace(",", ""))
    except ValueError:
        raise HTTPException(status_code=400, detail="quantity는 숫자여야 합니다.")
    success = update_inbound(int(row_index), date, item_name, qty, memo, verify_item)
    if not success:
        raise HTTPException(status_code=500, detail="입고 기록 수정에 실패했습니다.")
    return {"status": "success"}


@router.delete("/inbound")
def remove_inbound(
    row_index: int = Query(..., description="삭제할 행 번호"),
    item_name: str = Query(..., description="품목명 검증용"),
):
    """입고 기록 삭제"""
    success = delete_inbound(row_index, item_name)
    if not success:
        raise HTTPException(status_code=500, detail="입고 기록 삭제에 실패했습니다.")
    return {"status": "success"}


@router.get("/inventory-report")
def inventory_report():
    """품목별 + 월별 입출고 현황 리포트 반환"""
    return get_inventory_report()


# ── 실재고 직접 수정 ──────────────────────────────────────────────

@router.patch("/items/stock")
def update_item_stock(data: Dict[str, Any] = Body(...)):
    """토너 실재고를 지정한 값으로 직접 설정 (상세 수정 폼에서 사용)"""
    item_name = (data.get("item_name") or "").strip()
    new_stock  = data.get("stock")
    if not item_name or new_stock is None:
        raise HTTPException(status_code=400, detail="item_name과 stock이 필요합니다.")
    try:
        stock_int = int(str(new_stock).replace(',', ''))
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="stock은 정수여야 합니다.")
    success = set_toner_stock_direct(item_name, stock_int)
    if not success:
        raise HTTPException(status_code=500, detail="실재고 업데이트 실패 — 토너 시트에서 품목을 확인하세요.")
    return {"status": "success"}


# ── 개별 입고 내역 ──────────────────────────────────────────────

@router.get("/individual-inbound")
def list_individual_inbound(month: str = Query(None, description="조회할 월 (예: '2026년5월'), 없으면 전체")):
    """개별 입고 내역 조회"""
    return get_individual_inbound_history(month=month)


@router.post("/individual-inbound")
def create_individual_inbound(data: Dict[str, Any] = Body(...)):
    """개별 입고 추가 (실재고 자동 반영)"""
    if not data.get("item_name"):
        raise HTTPException(status_code=400, detail="품목명은 필수입니다.")
    if not data.get("date"):
        raise HTTPException(status_code=400, detail="날짜는 필수입니다.")
    qty = data.get("quantity", 0)
    if not qty or int(str(qty).replace(',', '')) <= 0:
        raise HTTPException(status_code=400, detail="수량은 1 이상이어야 합니다.")
    success = add_individual_inbound(data)
    if not success:
        raise HTTPException(status_code=500, detail="개별 입고 추가 실패")
    return {"status": "success"}


@router.delete("/individual-inbound")
def remove_individual_inbound(
    row_index: int = Query(..., description="삭제할 행 번호"),
    item_name: str = Query("", description="품목명 (검증용)"),
    quantity:  int = Query(0,  description="수량 (실재고 차감용)"),
):
    """개별 입고 삭제 (실재고 차감)"""
    result = delete_individual_inbound(row_index, item_name, quantity)
    if not result.get("success"):
        raise HTTPException(status_code=500, detail=result.get("error", "삭제 실패"))
    return {"status": "success"}
